"""Excel, CSV, grafik ve CBS çıktıları."""

from __future__ import annotations

import io
import hashlib
import json
import math
import tempfile
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import geopandas as gpd
import numpy as np
import pandas as pd
import rasterio
import requests
from rasterio.io import MemoryFile
from rasterio.features import rasterize, shapes
from rasterio.plot import plotting_extent
from rasterio.transform import from_bounds
from PIL import Image
from scipy.spatial import cKDTree
from shapely.geometry import shape as shapely_shape
from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.figure import Figure
from matplotlib import dates as mdates
from matplotlib.cm import ScalarMappable
from matplotlib.colors import BoundaryNorm, LinearSegmentedColormap, Normalize
from matplotlib.patches import Patch, Rectangle
from matplotlib.ticker import FuncFormatter
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from zetriklim.spi import classify_spi
from zetriklim.map_styles import map_visual_style


NAVY = "063447"
TEAL = "00A6A6"
PALE = "E2F4F2"
AMBER = "FFAD33"


def _nice_scale_length(width_m: float) -> float:
    target = max(width_m / 3, 1)
    power = 10 ** np.floor(np.log10(target))
    return float(max(value * power for value in (1, 2, 5) if value * power <= target))


def _add_north_arrow(ax) -> None:
    ax.annotate(
        "K",
        xy=(0.94, 0.94),
        xytext=(0.94, 0.80),
        xycoords="axes fraction",
        ha="center",
        va="center",
        fontsize=12,
        fontweight="bold",
        color="#063447",
        arrowprops={"facecolor": "#071f2b", "edgecolor": "white", "linewidth": 0.8, "width": 4, "headwidth": 13},
        bbox={"facecolor": "white", "edgecolor": "#071f2b", "alpha": 1, "boxstyle": "round,pad=0.28"},
        zorder=10,
    )


def _add_scale_bar(ax, *, projected: bool = True) -> None:
    xmin, xmax = ax.get_xlim()
    ymin, ymax = ax.get_ylim()
    if projected:
        length = _nice_scale_length(abs(xmax - xmin))
        label = f"{length / 1000:g} km" if length >= 1000 else f"{length:g} m"
    else:
        # Coğrafi rasterlarda orta enlem için yaklaşık metre/derece dönüşümü.
        latitude = (ymin + ymax) / 2
        metres_per_degree = 111_320 * max(np.cos(np.deg2rad(latitude)), 0.15)
        length_m = _nice_scale_length(abs(xmax - xmin) * metres_per_degree)
        length = length_m / metres_per_degree
        label = f"{length_m / 1000:g} km" if length_m >= 1000 else f"{length_m:g} m"
    span_x, span_y = abs(xmax - xmin), abs(ymax - ymin)
    x0 = xmin + span_x * 0.055
    y0 = ymin + span_y * 0.055
    height = span_y * 0.020
    segments = 5
    segment_length = length / segments
    background = Rectangle(
        (x0 - span_x * 0.018, y0 - span_y * 0.035),
        length + span_x * 0.036,
        span_y * 0.12,
        facecolor="white",
        edgecolor="#263238",
        linewidth=0.8,
        zorder=9,
    )
    ax.add_patch(background)
    for index in range(segments):
        ax.add_patch(
            Rectangle(
                (x0 + index * segment_length, y0),
                segment_length,
                height,
                facecolor="#111111" if index % 2 == 0 else "white",
                edgecolor="#111111",
                linewidth=1.1,
                zorder=11,
            )
        )
    ax.text(x0, y0 - span_y * 0.014, "0", ha="center", va="top", fontsize=8.5, color="#111111", zorder=12)
    ax.text(x0 + length / 2, y0 - span_y * 0.014, f"{float(label.split()[0]) / 2:g}", ha="center", va="top", fontsize=8.5, color="#111111", zorder=12)
    ax.text(x0 + length, y0 - span_y * 0.014, label, ha="center", va="top", fontsize=9.5, fontweight="bold", color="#111111", zorder=12)
    ax.text(x0, y0 + height + span_y * 0.014, "GRAFİK ÖLÇEK", ha="left", va="bottom", fontsize=8.5, fontweight="bold", color="#263238", zorder=12)


def dataframe_to_csv(data: pd.DataFrame) -> bytes:
    return data.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def build_excel(
    data: pd.DataFrame,
    *,
    metadata_rows: list[tuple[str, object]],
    area_rows: list[tuple[str, object]],
    analysis_tables: dict[str, pd.DataFrame] | None = None,
) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        data.to_excel(writer, sheet_name="İklim Verisi", index=False)
        pd.DataFrame(metadata_rows, columns=["Alan", "Değer"]).to_excel(writer, sheet_name="Kaynak ve Yöntem", index=False)
        pd.DataFrame(area_rows, columns=["Alan Özelliği", "Değer"]).to_excel(writer, sheet_name="Çalışma Alanı", index=False)
        data.describe(include="all").transpose().reset_index().to_excel(writer, sheet_name="Veri Özeti", index=False)
        for name, table in (analysis_tables or {}).items():
            table.to_excel(writer, sheet_name=name[:31], index=False)

        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = PatternFill("solid", fgColor=NAVY)
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for column in sheet.columns:
                letter = column[0].column_letter
                width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
                sheet.column_dimensions[letter].width = max(width, 12)

        data_sheet = writer.book["İklim Verisi"]
        data_sheet.sheet_view.showGridLines = False
        for row in range(2, data_sheet.max_row + 1):
            if row % 2 == 0:
                for cell in data_sheet[row]:
                    cell.fill = PatternFill("solid", fgColor="F2FAF9")
        if data_sheet.max_column >= 2 and data_sheet.max_row >= 3:
            chart = LineChart()
            chart.title = "İklim Zaman Serisi"
            chart.style = 13
            chart.y_axis.title = "Değer"
            chart.x_axis.title = "Tarih"
            chart.add_data(
                Reference(data_sheet, min_col=5, max_col=min(data_sheet.max_column, 8), min_row=1, max_row=data_sheet.max_row),
                titles_from_data=True,
            )
            chart.set_categories(Reference(data_sheet, min_col=1, min_row=2, max_row=data_sheet.max_row))
            chart.height = 9
            chart.width = 18
            summary_sheet = writer.book["Veri Özeti"]
            summary_sheet.add_chart(chart, "H2")
    return output.getvalue()


def build_spi_thesis_excel(spi_table: pd.DataFrame) -> bytes:
    """Tezde kullanılan yedi sütunlu su-yılı SPI tablosunu ayrı dosya olarak üretir."""
    if spi_table is None or spi_table.empty or "Tarih" not in spi_table:
        raise ValueError("Tez biçimli SPI tablosu için tarihli SPI sonuçları gereklidir.")
    frame = spi_table.copy()
    frame["Tarih"] = pd.to_datetime(frame["Tarih"], errors="coerce")
    frame = frame.dropna(subset=["Tarih"]).sort_values("Tarih")
    if frame.empty:
        raise ValueError("Tez biçimli SPI tablosunda geçerli tarih bulunamadı.")
    frame["Su yılı başlangıcı"] = np.where(
        frame["Tarih"].dt.month >= 10,
        frame["Tarih"].dt.year,
        frame["Tarih"].dt.year - 1,
    )
    # İlk/son kısmi su yılları bilimsel tablodaki yıllık karşılaştırmayı bozmasın.
    counts = frame.groupby("Su yılı başlangıcı")["Tarih"].transform("count")
    frame = frame[counts == 12]
    terminal = frame.groupby("Su yılı başlangıcı", as_index=False).tail(1)

    rows: list[dict[str, object]] = []
    for _, item in terminal.iterrows():
        start_year = int(item["Su yılı başlangıcı"])
        row: dict[str, object] = {"Yıllar": f"{start_year}-{start_year + 1}"}
        for scale in (3, 6, 12):
            value = pd.to_numeric(pd.Series([item.get(f"SPI-{scale}")]), errors="coerce").iloc[0]
            row[f"SPI İndis Değerleri {scale} Ay"] = (
                round(float(value), 2) if pd.notna(value) else None
            )
            row[f"Sınıflandırma {scale} Ay"] = (
                classify_spi(float(value)) if pd.notna(value) else "Hesaplanamadı"
            )
        rows.append(row)
    result = pd.DataFrame(rows)
    columns = [
        "Yıllar",
        "SPI İndis Değerleri 3 Ay",
        "Sınıflandırma 3 Ay",
        "SPI İndis Değerleri 6 Ay",
        "Sınıflandırma 6 Ay",
        "SPI İndis Değerleri 12 Ay",
        "Sınıflandırma 12 Ay",
    ]
    result = result.reindex(columns=columns)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="Sayfa1", index=False)
        sheet = writer.book["Sayfa1"]
        for column in (3, 5, 7):
            sheet.cell(1, column).value = "Sınıflandırma"
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = True
        widths = {"A": 15, "B": 25, "C": 22, "D": 25, "E": 22, "F": 27, "G": 22}
        for letter, width in widths.items():
            sheet.column_dimensions[letter].width = width
        sheet.row_dimensions[1].height = 34
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        class_fills = {
            "Aşırı kurak": "C65911",
            "Şiddetli kurak": "ED7D31",
            "Orta kurak": "F4B183",
            "Normale yakın": "FFFFFF",
            "Orta nemli": "BDD7EE",
            "Çok nemli": "9DC3E6",
            "Aşırı nemli": "4472C4",
            "Hesaplanamadı": "E7E6E6",
        }
        for row_number in range(2, sheet.max_row + 1):
            for value_column in (2, 4, 6):
                sheet.cell(row_number, value_column).number_format = "0.00"
            for class_column in (3, 5, 7):
                cell = sheet.cell(row_number, class_column)
                color = class_fills.get(str(cell.value), "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)
                if color in {"C65911", "4472C4"}:
                    cell.font = Font(color="FFFFFF")
    return output.getvalue()


def _figure_png(fig: Figure) -> bytes:
    fig.tight_layout()
    output = io.BytesIO()
    FigureCanvasAgg(fig).print_png(output)
    return output.getvalue()


def _monthly_numeric_frame(data: pd.DataFrame) -> pd.DataFrame:
    frame = data.copy()
    frame["Tarih"] = pd.to_datetime(frame["Tarih"], errors="coerce")
    frame = frame.dropna(subset=["Tarih"]).set_index("Tarih").sort_index()
    numeric = frame.apply(lambda values: pd.to_numeric(values, errors="coerce"))
    aggregations = {}
    for column in numeric.columns:
        label = str(column).casefold()
        aggregations[column] = "sum" if any(
            token in label for token in ("yağış", "precip", "pet", "et₀", "evapotrans")
        ) else "mean"
    return numeric.resample("MS").agg(aggregations)


def build_academic_chart_suite(
    data: pd.DataFrame,
    *,
    drought_table: pd.DataFrame | None = None,
    event_table: pd.DataFrame | None = None,
) -> dict[str, bytes]:
    """Her biri ayrı indirilebilen, yayın kalitesinde tamamlayıcı grafikler üretir."""
    monthly = _monthly_numeric_frame(data)
    charts: dict[str, bytes] = {}
    rain_columns = [c for c in monthly if any(t in c.casefold() for t in ("yağış", "precip"))]
    pet_columns = [c for c in monthly if any(t in c.casefold() for t in ("pet", "et₀", "evapotrans"))]
    temp_columns = [c for c in monthly if any(t in c.casefold() for t in ("sıcak", "temperature"))]
    rain = monthly[rain_columns[0]] if rain_columns else None
    pet = monthly[pet_columns[0]] if pet_columns else None
    mean_temp_columns = [c for c in temp_columns if "ortalama" in c.casefold() or "mean" in c.casefold()]
    mean_temp = monthly[mean_temp_columns[0] if mean_temp_columns else temp_columns[0]] if temp_columns else None

    def base_figure(rows: int = 1, height: float = 4.8):
        fig = Figure(figsize=(12, height), dpi=180, facecolor="#f7f5ee")
        axes = fig.subplots(rows, 1, sharex=rows > 1, squeeze=False).ravel()
        for ax in axes:
            ax.set_facecolor("#ffffff")
            ax.grid(alpha=0.24, linestyle=":", color="#78909c")
        return fig, axes

    if rain is not None and rain.notna().any():
        fig, (ax,) = base_figure()
        ax.bar(rain.index, rain, width=23, color="#2c7fb8", alpha=0.64, label="Aylık toplam")
        ax.plot(rain.index, rain.rolling(12, min_periods=6).mean(), color="#08306b", linewidth=2.1, label="12 aylık hareketli ortalama")
        ax.set_title("Aylık Yağış Zaman Serisi", color="#063447", weight="bold")
        ax.set_ylabel("Yağış (mm)")
        ax.legend(ncol=2, fontsize=8)
        ax.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=6, maxticks=12))
        ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(ax.xaxis.get_major_locator()))
        charts["grafik-01-yagis-zaman-serisi.png"] = _figure_png(fig)

    if mean_temp is not None and mean_temp.notna().any():
        fig, (ax,) = base_figure()
        min_cols = [c for c in temp_columns if "minimum" in c.casefold() or "min" in c.casefold()]
        max_cols = [c for c in temp_columns if "maksimum" in c.casefold() or "max" in c.casefold()]
        if min_cols and max_cols:
            ax.fill_between(monthly.index, monthly[min_cols[0]].to_numpy(float), monthly[max_cols[0]].to_numpy(float), color="#fdae61", alpha=0.25, label="Minimum–maksimum aralığı")
        ax.plot(mean_temp.index, mean_temp, color="#d95f0e", linewidth=0.9, alpha=0.65, label="Aylık ortalama")
        ax.plot(mean_temp.index, mean_temp.rolling(12, min_periods=6).mean(), color="#8c2d04", linewidth=2.1, label="12 aylık hareketli ortalama")
        ax.set_title("Aylık Hava Sıcaklığı Zaman Serisi", color="#063447", weight="bold")
        ax.set_ylabel("Sıcaklık (°C)")
        ax.legend(ncol=2, fontsize=8)
        ax.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=6, maxticks=12))
        ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(ax.xaxis.get_major_locator()))
        charts["grafik-02-sicaklik-zaman-serisi.png"] = _figure_png(fig)

    if rain is not None and pet is not None and rain.notna().any() and pet.notna().any():
        balance = rain - pet
        fig, axes = base_figure(2, 7.0)
        axes[0].plot(rain.index, rain, color="#2166ac", linewidth=1.0, label="Yağış")
        axes[0].plot(pet.index, pet, color="#e66101", linewidth=1.0, label="PET / ET₀")
        axes[0].set_ylabel("Su yüksekliği (mm)")
        axes[0].legend(ncol=2, fontsize=8)
        colors = np.where(balance >= 0, "#1b9e77", "#d73027")
        axes[1].bar(balance.index, balance, width=23, color=colors, alpha=0.78)
        axes[1].axhline(0, color="#37474f", linewidth=0.8)
        axes[1].set_ylabel("P − PET (mm)")
        axes[1].set_xlabel("Tarih")
        axes[0].set_title("Aylık Yağış, Potansiyel Evapotranspirasyon ve Su Dengesi", color="#063447", weight="bold")
        axes[1].xaxis.set_major_locator(mdates.AutoDateLocator(minticks=6, maxticks=12))
        axes[1].xaxis.set_major_formatter(mdates.ConciseDateFormatter(axes[1].xaxis.get_major_locator()))
        charts["grafik-03-su-dengesi.png"] = _figure_png(fig)

    drought = None
    drought_columns: list[str] = []
    if drought_table is not None and not drought_table.empty and "Tarih" in drought_table:
        drought = drought_table.copy()
        drought["Tarih"] = pd.to_datetime(drought["Tarih"], errors="coerce")
        drought = drought.dropna(subset=["Tarih"]).set_index("Tarih").sort_index()
        drought_columns = [
            c for c in drought.columns
            if str(c).upper().startswith(("SPI-", "SPEI-")) and "sınıf" not in str(c).casefold()
        ]
    if drought is not None and drought_columns:
        palette = ["#1b9e77", "#d95f02", "#7570b3", "#e7298a", "#66a61e", "#1f78b4"]

        def add_index_figure(prefix: str, file_name: str, title: str) -> None:
            columns = [
                column for column in drought_columns
                if str(column).upper().startswith(f"{prefix}-")
            ]
            if not columns:
                return
            fig, (ax,) = base_figure(height=5.2)
            for index, column in enumerate(columns[:8]):
                values = pd.to_numeric(drought[column], errors="coerce")
                ax.plot(
                    drought.index,
                    values,
                    linewidth=1.05,
                    color=palette[index % len(palette)],
                    label=column,
                )
            ax.axhspan(-3, -1, color="#d73027", alpha=0.09, label="Kuraklık eşiği (≤ −1)")
            ax.axhline(-1, color="#b71c1c", linestyle="--", linewidth=1)
            ax.axhline(0, color="#455a64", linewidth=0.8)
            ax.axhline(1, color="#1565c0", linestyle="--", linewidth=1)
            ax.set_title(title, color="#063447", weight="bold")
            ax.set_ylabel(f"{prefix} · standartlaştırılmış indis")
            ax.legend(ncol=3, fontsize=7.5)
            ax.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=6, maxticks=12))
            ax.xaxis.set_major_formatter(
                mdates.ConciseDateFormatter(ax.xaxis.get_major_locator())
            )
            charts[file_name] = _figure_png(fig)

        add_index_figure(
            "SPI",
            "grafik-04-spi-serileri.png",
            "Çok Ölçekli Standartlaştırılmış Yağış İndisi (SPI)",
        )
        add_index_figure(
            "SPEI",
            "grafik-05-spei-serileri.png",
            "Çok Ölçekli Standartlaştırılmış Yağış–Evapotranspirasyon İndisi (SPEI)",
        )

        if rain is not None and rain.notna().any():
            preferred = next((c for c in drought_columns if c.upper() == "SPI-3"), drought_columns[0])
            joined = pd.concat([rain.rename("Yağış"), pd.to_numeric(drought[preferred], errors="coerce").rename(preferred)], axis=1)
            fig, axes = base_figure(2, 7.0)
            axes[0].bar(joined.index, joined["Yağış"], width=23, color="#3182bd", alpha=0.72)
            axes[0].set_ylabel("Yağış (mm)")
            index_values = joined[preferred]
            axes[1].fill_between(joined.index, 0, index_values.to_numpy(float), where=(index_values.to_numpy(float) < 0), color="#d73027", alpha=0.55, interpolate=True)
            axes[1].fill_between(joined.index, 0, index_values.to_numpy(float), where=(index_values.to_numpy(float) >= 0), color="#2c7bb6", alpha=0.45, interpolate=True)
            axes[1].plot(joined.index, index_values, color="#263238", linewidth=0.8)
            axes[1].axhline(-1, color="#b71c1c", linestyle="--", linewidth=1)
            axes[1].set_ylabel(preferred)
            axes[1].set_xlabel("Tarih")
            axes[0].set_title(f"Yağış ve Kuraklık İlişkisi ({preferred})", color="#063447", weight="bold")
            axes[1].xaxis.set_major_locator(mdates.AutoDateLocator(minticks=6, maxticks=12))
            axes[1].xaxis.set_major_formatter(mdates.ConciseDateFormatter(axes[1].xaxis.get_major_locator()))
            charts["grafik-06-yagis-kuraklik-iliskisi.png"] = _figure_png(fig)

    if rain is not None or mean_temp is not None:
        fig, axes = base_figure(2 if rain is not None and mean_temp is not None else 1, 6.8)
        month_labels = ["Oca", "Şub", "Mar", "Nis", "May", "Haz", "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"]
        axis_index = 0
        if rain is not None:
            climatology = rain.groupby(rain.index.month).mean().reindex(range(1, 13))
            axes[axis_index].bar(range(1, 13), climatology, color="#2c7fb8", alpha=0.76)
            axes[axis_index].set_ylabel("Ortalama yağış (mm)")
            axis_index += 1
        if mean_temp is not None:
            climatology = mean_temp.groupby(mean_temp.index.month).mean().reindex(range(1, 13))
            axes[axis_index].plot(range(1, 13), climatology, marker="o", color="#d95f0e", linewidth=2)
            axes[axis_index].set_ylabel("Ortalama sıcaklık (°C)")
        for ax in axes:
            ax.set_xticks(range(1, 13), month_labels)
        axes[0].set_title("Aylık Klimatoloji ve Mevsimsel Döngü", color="#063447", weight="bold")
        axes[-1].set_xlabel("Takvim ayı")
        charts["grafik-07-mevsimsel-klimatoloji.png"] = _figure_png(fig)

    if event_table is not None and not event_table.empty and {"Başlangıç", "Bitiş"}.issubset(event_table.columns):
        events = event_table.copy()
        events["Başlangıç"] = pd.to_datetime(events["Başlangıç"], errors="coerce")
        events["Bitiş"] = pd.to_datetime(events["Bitiş"], errors="coerce")
        events = events.dropna(subset=["Başlangıç", "Bitiş"]).sort_values("Başlangıç").tail(40)
        if not events.empty:
            fig, (ax,) = base_figure(height=max(5.0, min(10.0, 2.5 + len(events) * 0.18)))
            for row, (_, event) in enumerate(events.iterrows()):
                start_num = mdates.date2num(event["Başlangıç"])
                width = max(mdates.date2num(event["Bitiş"]) - start_num, 1)
                severity = float(pd.to_numeric(pd.Series([event.get("Şiddet", 1)]), errors="coerce").fillna(1).iloc[0])
                ax.barh(row, width, left=start_num, height=0.68, color="#b2182b", alpha=min(0.45 + severity / 12, 0.9))
            ax.set_yticks(range(len(events)), [f"Olay {i + 1}" for i in range(len(events))], fontsize=7)
            ax.xaxis_date()
            ax.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=6, maxticks=12))
            ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(ax.xaxis.get_major_locator()))
            ax.set_title("Kuraklık Olaylarının Zaman İçindeki Süreleri", color="#063447", weight="bold")
            ax.set_xlabel("Başlangıç–bitiş aralığı")
            charts["grafik-08-kuraklik-olay-takvimi.png"] = _figure_png(fig)

    return charts


def build_raster_png(
    geotiff: bytes,
    title: str,
    *,
    boundary: gpd.GeoDataFrame | None = None,
    palette: str = "RdBu",
    colorbar_label: str = "SPI",
    fixed_range: tuple[float, float] | None = (-2.5, 2.5),
) -> bytes:
    with MemoryFile(geotiff) as memory:
        with memory.open() as src:
            raster = src.read(1, masked=True).astype(float).filled(float("nan"))
            extent = plotting_extent(src)
            raster_crs = src.crs
    fig = Figure(figsize=(9, 7), dpi=150, facecolor="#f7f5ee")
    ax = fig.add_subplot(111)
    range_args = (
        {"vmin": fixed_range[0], "vmax": fixed_range[1]}
        if fixed_range is not None
        else {}
    )
    image = ax.imshow(raster, cmap=palette, extent=extent, origin="upper", **range_args)
    if boundary is not None and raster_crs is not None:
        boundary.to_crs(raster_crs).boundary.plot(
            ax=ax,
            color="#082f49",
            linewidth=1.4,
            zorder=3,
        )
    ax.set_title(title, color="#063447", fontsize=13, weight="bold", pad=12)
    projected = bool(raster_crs and raster_crs.is_projected)
    ax.set_xlabel("Doğu (m)" if projected else "Boylam (°)")
    ax.set_ylabel("Kuzey (m)" if projected else "Enlem (°)")
    ax.grid(color="#607d8b", linestyle=":", linewidth=0.55, alpha=0.45)
    _add_north_arrow(ax)
    _add_scale_bar(ax, projected=projected)
    colorbar = fig.colorbar(image, ax=ax, shrink=0.78)
    colorbar.set_label(colorbar_label)
    fig.tight_layout()
    output = io.BytesIO()
    FigureCanvasAgg(fig).print_png(output)
    return output.getvalue()


def _lonlat_to_tile(longitude: float, latitude: float, zoom: int) -> tuple[int, int]:
    latitude = float(np.clip(latitude, -85.05112878, 85.05112878))
    count = 2**zoom
    x = int(np.floor((longitude + 180.0) / 360.0 * count))
    latitude_rad = math.radians(latitude)
    y = int(
        np.floor(
            (1.0 - math.asinh(math.tan(latitude_rad)) / math.pi)
            / 2.0
            * count
        )
    )
    return int(np.clip(x, 0, count - 1)), int(np.clip(y, 0, count - 1))


def _tile_grid(bounds: list[float], *, maximum_tiles: int = 4) -> tuple[int, range, range]:
    """Küçük ve güvenilir bir web-karo ızgarası seçer.

    Dört karo sınırı, tek karonun bazı alanlarda oluşturduğu yaklaşık 1.2 km'lik
    çıktı ızgarasını çoğunlukla yaklaşık 0.6 km'ye indirirken eski 16 karolu
    isteğin bağlantı hatası riskine geri dönmemeyi sağlar.
    """
    west, south, east, north = map(float, bounds)
    for zoom in range(12, 5, -1):
        x_min, y_max = _lonlat_to_tile(west, south, zoom)
        x_max, y_min = _lonlat_to_tile(east, north, zoom)
        xs = range(min(x_min, x_max), max(x_min, x_max) + 1)
        ys = range(min(y_min, y_max), max(y_min, y_max) + 1)
        if len(xs) * len(ys) <= maximum_tiles:
            return zoom, xs, ys
    x_min, y_max = _lonlat_to_tile(west, south, 6)
    x_max, y_min = _lonlat_to_tile(east, north, 6)
    return 6, range(min(x_min, x_max), max(x_min, x_max) + 1), range(min(y_min, y_max), max(y_min, y_max) + 1)


def _download_map_tile(
    tile_url: str, zoom: int, x: int, y: int
) -> tuple[Image.Image, float, str]:
    url = tile_url.replace("{z}", str(zoom)).replace("{x}", str(x)).replace("{y}", str(y))
    started = time.perf_counter()
    last_error: Exception | None = None
    retry_delays = (0.0, 0.75, 1.5, 3.0, 6.0)
    for attempt, retry_delay in enumerate(retry_delays):
        if retry_delay:
            time.sleep(retry_delay)
        try:
            response = requests.get(
                url,
                timeout=12,
                headers={"Cache-Control": "no-cache", "Pragma": "no-cache"},
            )
            response.raise_for_status()
            tile = Image.open(io.BytesIO(response.content)).convert("RGBA")
            if tile.size != (256, 256):
                tile = tile.resize((256, 256), Image.Resampling.BILINEAR)
            return (
                tile,
                time.perf_counter() - started,
                hashlib.sha256(tile.tobytes()).hexdigest(),
            )
        except Exception as error:  # Ağ ve geçici karo hataları aynı yeniden deneme akışındadır.
            last_error = error
    raise RuntimeError(f"Harita karosu indirilemedi: {last_error}")


def _rendered_data_mask(rgba: np.ndarray) -> np.ndarray:
    """MapID PNG içindeki gerçek veri piksellerini ayırır.

    Climate Engine bazı başarısız/henüz hazır olmayan karo parçalarını HTTP 200
    ve opak saf beyaz piksellerle döndürebiliyor. Yalnız alfa kanalına bakmak bu
    parçaları veri sanıp %100 kapsam raporluyordu. Kullanılan paletlerdeki en açık
    geçerli renkler saf beyaz değildir; bu nedenle yalnız nötr, neredeyse saf beyaz
    sunucu arka planı NoData kabul edilir.
    """
    rgb = rgba[..., :3].astype(np.int16)
    near_white = (rgb.min(axis=2) >= 252) & ((rgb.max(axis=2) - rgb.min(axis=2)) <= 3)
    return (rgba[..., 3] > 0) & ~near_white


def _fill_from_parent_tiles(
    array: np.ndarray,
    missing: np.ndarray,
    tile_url: str,
    zoom: int,
    xs: range,
    ys: range,
    *,
    maximum_levels: int = 2,
) -> tuple[np.ndarray, np.ndarray, int, list[str]]:
    """Eksik hedef pikselleri aynı MapID'nin daha düşük zoom karolarıyla tamamlar."""
    filled = array.copy()
    remaining = missing.copy()
    fallback_tiles = 0
    errors: list[str] = []
    for level in range(1, maximum_levels + 1):
        if not remaining.any() or zoom - level < 0:
            break
        parent_zoom = zoom - level
        divisor = 2**level
        parents: dict[tuple[int, int], Image.Image] = {}
        needed = {
            (x // divisor, y // divisor)
            for y in ys
            for x in xs
            if remaining[
                (y - ys.start) * 256:(y - ys.start + 1) * 256,
                (x - xs.start) * 256:(x - xs.start + 1) * 256,
            ].any()
        }
        with ThreadPoolExecutor(max_workers=min(4, len(needed) or 1)) as executor:
            futures = {
                executor.submit(_download_map_tile, tile_url, parent_zoom, x, y): (x, y)
                for x, y in needed
            }
            for future in as_completed(futures):
                parent_x, parent_y = futures[future]
                try:
                    parents[(parent_x, parent_y)] = future.result()[0]
                    fallback_tiles += 1
                except Exception as error:
                    errors.append(f"z{parent_zoom}/{parent_x}/{parent_y}: {error}")
        for y in ys:
            for x in xs:
                parent = parents.get((x // divisor, y // divisor))
                if parent is None:
                    continue
                local_x = x % divisor
                local_y = y % divisor
                left = round(local_x * 256 / divisor)
                top = round(local_y * 256 / divisor)
                right = round((local_x + 1) * 256 / divisor)
                bottom = round((local_y + 1) * 256 / divisor)
                candidate = np.asarray(
                    parent.crop((left, top, right, bottom)).resize(
                        (256, 256), Image.Resampling.BILINEAR
                    ).convert("RGBA")
                )
                target_rows = slice((y - ys.start) * 256, (y - ys.start + 1) * 256)
                target_cols = slice((x - xs.start) * 256, (x - xs.start + 1) * 256)
                target_missing = remaining[target_rows, target_cols]
                candidate_valid = _rendered_data_mask(candidate)
                use = target_missing & candidate_valid
                if use.any():
                    target = filled[target_rows, target_cols]
                    target[use] = candidate[use]
                    remaining[target_rows, target_cols][use] = False
    return filled, remaining, fallback_tiles, errors


def _decode_rendered_values(
    rgba: np.ndarray,
    style: dict[str, object],
) -> np.ndarray:
    """Climate Engine renk rampasını yaklaşık sayısal değerlere geri dönüştürür.

    MapID servisi analiz rasterını PNG olarak sunduğundan renk rampası terslenir.
    Çıktı tek bantlı sayısal CBS rasterıdır; RGB ekran görüntüsü değildir.
    """
    colors = [str(color) for color in style["colors"]]
    color_map = LinearSegmentedColormap.from_list("ce_decode", colors, N=2048)
    lut = np.rint(color_map(np.linspace(0, 1, 2048))[:, :3] * 255).astype(np.int16)
    valid = rgba[..., 3] > 0
    values = np.full(rgba.shape[:2], np.nan, dtype="float32")
    if not valid.any():
        return values
    pixels = rgba[..., :3][valid]
    unique, inverse = np.unique(pixels, axis=0, return_inverse=True)
    _, positions = cKDTree(lut.astype("float32")).query(unique.astype("float32"), k=1)
    minimum = float(style["minimum"])
    maximum = float(style["maximum"])
    decoded = minimum + positions.astype("float32") / (len(lut) - 1) * (maximum - minimum)
    values[valid] = decoded[inverse]
    return values


def _display_normalization(
    values: np.ndarray,
    analysis: str,
    style: dict[str, object],
):
    finite = values[np.isfinite(values)]
    colors = [str(color) for color in style["colors"]]
    color_map = LinearSegmentedColormap.from_list(f"{analysis}_map", colors, N=256)
    if style.get("index"):
        return color_map, BoundaryNorm(style["index"], color_map.N), (
            float(style["minimum"]), float(style["maximum"])
        )
    fixed_min = float(style["minimum"])
    fixed_max = float(style["maximum"])
    if not len(finite) or "Anomalisi" in analysis:
        return color_map, Normalize(fixed_min, fixed_max), (fixed_min, fixed_max)
    low, high = np.nanpercentile(finite, [2, 98])
    low, high = max(float(low), fixed_min), min(float(high), fixed_max)
    minimum_span = max((fixed_max - fixed_min) * 0.04, 0.05)
    if high - low < minimum_span:
        middle = (high + low) / 2
        low, high = middle - minimum_span / 2, middle + minimum_span / 2
    low, high = max(low, fixed_min), min(high, fixed_max)
    return color_map, Normalize(low, high), (low, high)


def _build_map_exports(
    values: np.ndarray,
    valid: np.ndarray,
    transform,
    geographic: gpd.GeoDataFrame,
    analysis: str,
    metadata: dict[str, object],
) -> tuple[bytes, bytes]:
    """Sayısal analiz değerlerini GeoTIFF ve sınıflı Shapefile olarak üretir."""
    height, width = values.shape
    nodata = -9999.0
    raster_values = np.where(valid & np.isfinite(values), values, nodata).astype("float32")
    profile = {
        "driver": "GTiff",
        "width": width,
        "height": height,
        "count": 1,
        "dtype": "float32",
        "crs": "EPSG:3857",
        "transform": transform,
        "compress": "deflate",
        "nodata": nodata,
    }
    with MemoryFile() as memory:
        with memory.open(**profile) as dataset:
            dataset.write(raster_values, 1)
            dataset.set_band_description(1, str(analysis))
            dataset.update_tags(**{key.upper(): str(value) for key, value in metadata.items()})
        geotiff_bytes = memory.read()

    style = map_visual_style(analysis) or {
        "minimum": 0.0,
        "maximum": 1.0,
        "colors": ["#2166ac", "#f7f7f7", "#b2182b"],
    }
    target_width = min(width, 160)
    target_height = max(1, round(height * target_width / max(width, 1)))
    target_height = min(target_height, 160)
    downsampled = np.asarray(Image.fromarray(raster_values, mode="F").resize(
        (target_width, target_height), Image.Resampling.NEAREST
    ), dtype="float32")
    down_valid = np.asarray(Image.fromarray(valid.astype("uint8") * 255).resize(
        (target_width, target_height), Image.Resampling.NEAREST
    )) > 0
    down_transform = transform * rasterio.Affine.scale(
        width / target_width,
        height / target_height,
    )
    if style.get("index") and len(style["index"]) == len(style["colors"]) + 1:
        class_edges = [float(value) for value in style["index"]]
    else:
        class_edges = np.linspace(
            float(style["minimum"]),
            float(style["maximum"]),
            len(style["colors"]) + 1,
        ).tolist()
    classified = np.digitize(downsampled, class_edges[1:-1], right=False).astype("uint8") + 1
    classified[~down_valid] = 0
    boundary = geographic.to_crs(3857).geometry.union_all()
    records: list[dict[str, object]] = []
    geometries = []
    for geometry_mapping, class_value in shapes(
        classified,
        mask=down_valid,
        transform=down_transform,
    ):
        class_index = int(class_value) - 1
        geometry = shapely_shape(geometry_mapping).intersection(boundary)
        if geometry.is_empty or class_index < 0:
            continue
        records.append(
            {
                "SINIF": class_index + 1,
                "ALT": round(class_edges[class_index], 4),
                "UST": round(class_edges[class_index + 1], 4),
                "RENK": str(style["colors"][class_index]),
                "ANALIZ": str(analysis)[:80],
            }
        )
        geometries.append(geometry)
    vector_layer = gpd.GeoDataFrame(records, geometry=geometries, crs=3857).to_crs(4326)

    with tempfile.TemporaryDirectory(prefix="zetriklim_map_") as temp:
        temp_path = Path(temp)
        shp_dir = temp_path / "shapefile"
        shp_dir.mkdir()
        shp_path = shp_dir / "analiz_siniflari.shp"
        vector_layer.to_file(shp_path, driver="ESRI Shapefile", encoding="UTF-8")
        (shp_dir / "harita-metadata.json").write_text(
            json.dumps(metadata, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        shp_output = io.BytesIO()
        with zipfile.ZipFile(shp_output, "w", zipfile.ZIP_DEFLATED) as archive:
            for path in sorted(shp_dir.iterdir()):
                archive.write(path, arcname=path.name)
        return geotiff_bytes, shp_output.getvalue()


def build_tile_map_png(
    tile_url: str,
    gdf: gpd.GeoDataFrame,
    *,
    title: str,
    analysis: str,
    source: str,
    period: str,
) -> tuple[bytes, dict[str, object]]:
    """Etkileşimli katmanın aynı karolarından, havzaya kırpılmış statik PNG üretir."""
    if not all(marker in tile_url for marker in ("{z}", "{x}", "{y}")):
        raise ValueError("Climate Engine karo adresi z/x/y şablonu içermiyor.")
    geographic = gdf.to_crs(4326)
    bounds = geographic.total_bounds.tolist()
    # En fazla dört karo, büyük alanlarda tek karodan daha sık bir çıktı ızgarası
    # sağlar. Bütün karoların eksiksiz gelmesi kritik iklim haritalarında aşağıda
    # hâlâ zorunludur; dolayısıyla çözünürlük artışı yarım harita pahasına yapılmaz.
    maximum_tiles = 4
    zoom, xs, ys = _tile_grid(bounds, maximum_tiles=maximum_tiles)
    tile_count = len(xs) * len(ys)
    mosaic = Image.new("RGBA", (len(xs) * 256, len(ys) * 256), (0, 0, 0, 0))
    downloaded = 0
    durations: list[float] = []
    tile_hashes: list[str] = []
    errors: list[str] = []
    # MapID adresi Earth Engine tarafından eşzamansız hazırlanabilir. Önce merkez
    # karoyu tek başına istemek, 4–16 karonun aynı anda 404/5xx ile tükenmesini
    # engeller ve sunucu hazır olduktan sonra kalan küçük grubu indirir.
    center_x = xs.start + len(xs) // 2
    center_y = ys.start + len(ys) // 2
    try:
        center_tile, center_duration, center_hash = _download_map_tile(
            tile_url, zoom, center_x, center_y
        )
        mosaic.paste(
            center_tile,
            ((center_x - xs.start) * 256, (center_y - ys.start) * 256),
        )
        downloaded = 1
        durations.append(center_duration)
        tile_hashes.append(center_hash)
    except Exception as error:
        errors.append(f"z{zoom}/{center_x}/{center_y}: {error}")
    remaining_tiles = [
        (x, y) for y in ys for x in xs if (x, y) != (center_x, center_y)
    ]
    with ThreadPoolExecutor(max_workers=min(2, len(remaining_tiles) or 1)) as executor:
        futures = {
            executor.submit(_download_map_tile, tile_url, zoom, x, y): (x, y)
            for x, y in remaining_tiles
        }
        for future in as_completed(futures):
            x, y = futures[future]
            try:
                tile, duration, tile_hash = future.result()
                mosaic.paste(tile, ((x - xs.start) * 256, (y - ys.start) * 256))
                downloaded += 1
                durations.append(duration)
                tile_hashes.append(tile_hash)
            except Exception as error:
                errors.append(f"z{zoom}/{x}/{y}: {error}")
    unique_tile_images = len(set(tile_hashes))
    repeated_tile_warning = (
        tile_count >= 4 and downloaded == tile_count and unique_tile_images <= 1
    )

    web_limit = 20_037_508.342789244
    world = 2 * web_limit
    count = 2**zoom
    x_min = -web_limit + xs.start / count * world
    x_max = -web_limit + xs.stop / count * world
    y_max = web_limit - ys.start / count * world
    y_min = web_limit - ys.stop / count * world
    array = np.asarray(mosaic).copy()
    transform = from_bounds(x_min, y_min, x_max, y_max, array.shape[1], array.shape[0])
    projected = geographic.to_crs(3857)
    mask = rasterize(
        [(geometry, 1) for geometry in projected.geometry if geometry is not None and not geometry.is_empty],
        out_shape=array.shape[:2],
        transform=transform,
        fill=0,
        default_value=1,
        dtype="uint8",
    )
    boundary_pixels = mask == 1
    primary_data = _rendered_data_mask(array)
    primary_valid_pixels = boundary_pixels & primary_data
    primary_coverage = float(primary_valid_pixels.sum() / max(boundary_pixels.sum(), 1))
    if analysis in {"SPI", "Yağış", "Sıcaklık"} and (
        downloaded != tile_count or primary_coverage < 0.995
    ):
        raise RuntimeError(
            "Birincil Climate Engine karosu çalışma alanını bütünüyle doldurmadı "
            f"({downloaded}/{tile_count} karo, %{primary_coverage * 100:.1f} piksel). "
            "Yanlış veya yarım harita üretmemek için yeni MapID istenmelidir."
        )
    fallback_tiles = 0
    fallback_errors: list[str] = []
    if primary_coverage < 0.999:
        array, _, fallback_tiles, fallback_errors = _fill_from_parent_tiles(
            array,
            boundary_pixels & ~primary_data,
            tile_url,
            zoom,
            xs,
            ys,
        )
    array[..., 3] = np.where(boundary_pixels, array[..., 3], 0)
    valid_pixels = boundary_pixels & _rendered_data_mask(array)
    coverage = float(valid_pixels.sum() / max(boundary_pixels.sum(), 1))
    if coverage < 0.999:
        raise RuntimeError(
            f"Harita eksik bırakılmadı: birincil kapsam %{primary_coverage * 100:.1f}, "
            f"çok çözünürlüklü tamamlama sonrası kapsam %{coverage * 100:.1f}. "
            "En az %99,9 gerçek piksel kapsamı sağlanamadığı için çıktı üretilmedi."
        )
    sampled = array[valid_pixels][:: max(1, int(valid_pixels.sum() / 25_000))]
    unique_colors = int(len(np.unique(sampled[:, :3], axis=0))) if len(sampled) else 0
    low_color_variation = unique_colors < 3
    quality = {
        "status": (
            "uygun"
            if coverage >= 0.999 and not repeated_tile_warning
            else "kontrol_gerekli"
        ),
        "zoom": zoom,
        "requested_tiles": tile_count,
        "downloaded_tiles": downloaded,
        "failed_tiles": tile_count - downloaded,
        "coverage_ratio": round(coverage, 4),
        "primary_coverage_ratio": round(primary_coverage, 4),
        "fallback_tiles": fallback_tiles,
        "recovered_primary_failures": tile_count - downloaded if coverage >= 0.999 else 0,
        "coverage_method": "piksel_icerigi_ve_cok_cozunurluklu_tamamlama",
        "sample_unique_colors": unique_colors,
        "unique_tile_images": unique_tile_images,
        "mean_tile_seconds": round(float(np.mean(durations)), 3) if durations else None,
        "maximum_tile_seconds": round(float(np.max(durations)), 3) if durations else None,
        "warnings": (
            (["Bütün karolar aynı görüntüyü içeriyor; mekânsal değişkenlik ayrıca kontrol edilmelidir."]
             if repeated_tile_warning else [])
            + (["Havza içinde çok düşük mekânsal değişkenlik var; bu durum tek başına veri hatası değildir."]
               if low_color_variation else [])
            + errors[:10]
            + fallback_errors[:10]
        ),
    }
    style = map_visual_style(analysis) or {
        "minimum": 0.0,
        "maximum": 1.0,
        "colors": ["#2166ac", "#f7f7f7", "#b2182b"],
        "unit": analysis,
    }
    values = _decode_rendered_values(array, style)
    values[~valid_pixels] = np.nan
    finite_values = values[np.isfinite(values)]
    if not len(finite_values):
        raise RuntimeError("Harita renkleri sayısal analiz değerlerine dönüştürülemedi.")
    if analysis == "SPI":
        saturation_fraction = float(np.mean(np.abs(finite_values) >= 2.95))
        spatial_std = float(np.nanstd(finite_values))
        if saturation_fraction >= 0.95 or (
            abs(float(np.nanmean(finite_values))) >= 2.75 and spatial_std < 0.08
        ):
            raise RuntimeError(
                "Climate Engine SPI rasterı ±3 sınırında doygun ve mekânsal olarak "
                "tekdüze döndü; bu çıktı akademik kalite kontrolünden geçmedi. "
                "Yanlış harita yayımlanmadı ve yeni MapID istenmelidir."
            )
    color_map, normalizer, display_range = _display_normalization(values, analysis, style)
    rendered = np.rint(color_map(normalizer(values)) * 255).astype("uint8")
    rendered[..., 3] = np.where(valid_pixels, 255, 0).astype("uint8")
    export_pixel_size_m = round(abs(float(transform.a)), 2)
    source_upper = source.upper()
    source_native_resolution_m = (
        4800.0 if "CHIRPS" in source_upper
        else 9600.0 if "ERA5_AG" in source_upper or "ERA5 AG" in source_upper
        else None
    )
    quality.update(
        {
            "value_min": round(float(np.nanmin(finite_values)), 4),
            "value_max": round(float(np.nanmax(finite_values)), 4),
            "value_mean": round(float(np.nanmean(finite_values)), 4),
            "value_std": round(float(np.nanstd(finite_values)), 4),
            "spi_saturation_fraction": (
                round(float(np.mean(np.abs(finite_values) >= 2.95)), 4)
                if analysis == "SPI" else None
            ),
            "display_min": round(float(display_range[0]), 4),
            "display_max": round(float(display_range[1]), 4),
            # pixel_size_m geriye dönük uyumluluk için korunur; bu değer kaynak
            # veri çözünürlüğü değil, indirilen/üretilen GeoTIFF ızgara aralığıdır.
            "pixel_size_m": export_pixel_size_m,
            "export_grid_size_m": export_pixel_size_m,
            "source_native_resolution_m": source_native_resolution_m,
            "effective_resolution_m": (
                max(export_pixel_size_m, source_native_resolution_m)
                if source_native_resolution_m is not None else export_pixel_size_m
            ),
            "qgis_raster_type": "float32_single_band",
            "qgis_nodata": -9999.0,
        }
    )
    if (
        source_native_resolution_m is not None
        and export_pixel_size_m < source_native_resolution_m
    ):
        quality["warnings"].append(
            "Çıktı ızgarası kaynak rasterdan daha sıktır; ara değerler yeni "
            "meteorolojik ayrıntı oluşturmaz. Bilimsel etkin çözünürlük kaynak "
            f"ürünün yaklaşık {source_native_resolution_m / 1000:.1f} km "
            "çözünürlüğüdür."
        )
    overlay_output = io.BytesIO()
    Image.fromarray(rendered, mode="RGBA").save(overlay_output, format="PNG", optimize=True)
    tile_count_axis = 2**zoom

    def tile_longitude(tile_x: int) -> float:
        return tile_x / tile_count_axis * 360.0 - 180.0

    def tile_latitude(tile_y: int) -> float:
        return math.degrees(math.atan(math.sinh(math.pi * (1 - 2 * tile_y / tile_count_axis))))

    quality["_overlay_png"] = overlay_output.getvalue()
    quality["_overlay_bounds"] = [
        [tile_latitude(ys.stop), tile_longitude(xs.start)],
        [tile_latitude(ys.start), tile_longitude(xs.stop)],
    ]
    geotiff_bytes, classified_shp_bytes = _build_map_exports(
        values,
        valid_pixels,
        transform,
        geographic,
        analysis,
        {
            "harita": title,
            "analiz": analysis,
            "kaynak": source,
            "donem": period,
            "kapsama": f"%{coverage * 100:.1f}",
            "deger_uretim_yontemi": (
                "Climate Engine MapID PNG renk rampasının 2048 adımlı terslenmesi; "
                "tek bant Float32"
            ),
        },
    )
    quality["_geotiff_bytes"] = geotiff_bytes
    quality["_classified_shp_bytes"] = classified_shp_bytes

    fig = Figure(figsize=(10, 8), dpi=180, facecolor="#f4f2eb")
    ax = fig.add_subplot(111)
    image = ax.imshow(
        values,
        cmap=color_map,
        norm=normalizer,
        extent=(x_min, x_max, y_min, y_max),
        origin="upper",
        interpolation="bilinear",
    )
    projected.boundary.plot(ax=ax, color="#082f49", linewidth=1.5, zorder=3)
    area_bounds = projected.total_bounds
    span_x, span_y = area_bounds[2] - area_bounds[0], area_bounds[3] - area_bounds[1]
    ax.set_xlim(area_bounds[0] - span_x * 0.07, area_bounds[2] + span_x * 0.07)
    ax.set_ylim(area_bounds[1] - span_y * 0.09, area_bounds[3] + span_y * 0.07)
    ax.set_title(title, color="#063447", fontsize=15, weight="bold", pad=14)
    earth_radius = 6_378_137.0

    def longitude_label(value: float, _: object) -> str:
        longitude = math.degrees(value / earth_radius)
        return f"{abs(longitude):.2f}°{'D' if longitude >= 0 else 'B'}"

    def latitude_label(value: float, _: object) -> str:
        latitude = math.degrees(2 * math.atan(math.exp(value / earth_radius)) - math.pi / 2)
        return f"{abs(latitude):.2f}°{'K' if latitude >= 0 else 'G'}"

    ax.xaxis.set_major_formatter(FuncFormatter(longitude_label))
    ax.yaxis.set_major_formatter(FuncFormatter(latitude_label))
    ax.set_xlabel("Boylam (coğrafi derece)")
    ax.set_ylabel("Enlem (coğrafi derece)")
    ax.grid(color="#78909c", linestyle=":", linewidth=0.55, alpha=0.5)
    ax.set_axisbelow(True)
    _add_north_arrow(ax)
    if style:
        colorbar = fig.colorbar(
            image,
            ax=ax,
            shrink=0.78,
            pad=0.025,
        )
        colorbar.set_label(str(style["unit"]))
    fig.text(0.02, 0.018, f"Kaynak: {source} · Dönem: {period}", fontsize=7.5, color="#455a64")
    fig.text(
        0.98,
        0.018,
        f"Karo doğrulama: {downloaded}/{tile_count} · Alan kapsama: %{coverage * 100:.1f}",
        ha="right",
        fontsize=7.5,
        color="#455a64",
    )
    fig.subplots_adjust(left=0.10, right=0.91, bottom=0.15, top=0.91)
    map_width = abs(ax.get_xlim()[1] - ax.get_xlim()[0])
    scale_length = _nice_scale_length(map_width)
    map_position = ax.get_position()
    scale_width = map_position.width * scale_length / map_width
    scale_axis = fig.add_axes([map_position.x0, 0.058, scale_width, 0.032])
    for segment in range(5):
        scale_axis.add_patch(
            Rectangle(
                (segment / 5, 0.34),
                1 / 5,
                0.38,
                facecolor="black" if segment % 2 == 0 else "white",
                edgecolor="black",
                linewidth=0.8,
            )
        )
    scale_axis.set_xlim(0, 1)
    scale_axis.set_ylim(0, 1)
    scale_axis.axis("off")
    scale_axis.text(0, 0.04, "0", fontsize=7, ha="left", va="bottom", color="#263238")
    scale_axis.text(
        1,
        0.04,
        f"{scale_length / 1000:g} km" if scale_length >= 1000 else f"{scale_length:g} m",
        fontsize=7,
        ha="right",
        va="bottom",
        color="#263238",
        weight="bold",
    )
    scale_axis.text(0, 0.82, "ÖLÇEK", fontsize=7, ha="left", va="bottom", color="#263238", weight="bold")
    output = io.BytesIO()
    FigureCanvasAgg(fig).print_png(output)
    return output.getvalue(), quality


def build_timeseries_png(data: pd.DataFrame) -> bytes:
    dates = pd.to_datetime(data["Tarih"], errors="coerce")
    numeric = data.drop(columns=["Tarih", "Örnek ID", "Enlem", "Boylam"], errors="ignore").apply(
        lambda values: pd.to_numeric(values, errors="coerce")
    )
    numeric = numeric.loc[:, numeric.notna().any()]
    groups = {
        "Yağış ve su dengesi (mm)": [c for c in numeric if any(t in c.lower() for t in ("yağış", "precip", "pet", "et₀", "su dengesi"))],
        "Sıcaklık (°C)": [c for c in numeric if any(t in c.lower() for t in ("sıcak", "temperature", "lst"))],
        "Standart indis / bitki örtüsü": [c for c in numeric if c.upper().startswith(("SPI", "SPEI", "NDVI", "EVI"))],
    }
    used = {column for columns in groups.values() for column in columns}
    remainder = [column for column in numeric if column not in used]
    if remainder:
        groups["Diğer değişkenler"] = remainder
    groups = {label: columns[:8] for label, columns in groups.items() if columns}
    if not groups:
        groups = {"Veri": []}
    fig = Figure(figsize=(12, max(4.5, 3.1 * len(groups))), dpi=150, facecolor="#f7f5ee")
    axes = fig.subplots(len(groups), 1, sharex=True, squeeze=False).ravel()
    palette = ["#1676b8", "#e86f16", "#269e45", "#9c4dcc", "#c13f3f", "#008c95"]
    for ax, (label, columns) in zip(axes, groups.items()):
        is_temperature = label.startswith("Sıcaklık")
        mean_columns = [column for column in columns if "ortalama" in column.casefold()]
        min_columns = [column for column in columns if "minimum" in column.casefold()]
        max_columns = [column for column in columns if "maksimum" in column.casefold()]
        if is_temperature and mean_columns and min_columns and max_columns:
            mean_col, min_col, max_col = mean_columns[0], min_columns[0], max_columns[0]
            ax.fill_between(
                dates,
                numeric[min_col].to_numpy(float),
                numeric[max_col].to_numpy(float),
                color="#78b7d0",
                alpha=0.24,
                label="Aylık minimum–maksimum aralığı",
            )
            ax.plot(dates, numeric[mean_col], color="#075b68", linewidth=0.9, alpha=0.65, label=mean_col)
            rolling = numeric[mean_col].rolling(12, min_periods=6, center=True).mean()
            ax.plot(dates, rolling, color="#c63d2f", linewidth=2.0, label="12 aylık ortalama sıcaklık")
        else:
            for index, column in enumerate(columns):
                color = palette[index % len(palette)]
                ax.plot(dates, numeric[column], linewidth=0.8, alpha=0.52, color=color, label=column)
                if len(data) >= 60:
                    rolling = numeric[column].rolling(12, min_periods=6, center=True).mean()
                    ax.plot(dates, rolling, linewidth=1.75, color=color, label=f"{column} · 12 aylık ort.")
        if label.startswith("Standart indis"):
            ax.axhline(0, color="#455a64", linewidth=0.8)
            ax.axhline(-1, color="#b71c1c", linewidth=0.8, linestyle="--", alpha=0.75)
            ax.axhline(1, color="#1565c0", linewidth=0.8, linestyle="--", alpha=0.75)
        ax.set_ylabel(label, fontsize=9)
        ax.grid(alpha=0.22, linestyle=":", color="#607d8b")
        ax.set_facecolor("#fbfcfa")
        if columns:
            ax.legend(loc="best", fontsize=7, ncol=2)
    axes[0].set_title("Zetriklim – Analiz Zaman Serisi", color="#063447", fontsize=14, weight="bold")
    axes[-1].set_xlabel("Tarih")
    axes[-1].xaxis.set_major_locator(mdates.AutoDateLocator(minticks=6, maxticks=12))
    axes[-1].xaxis.set_major_formatter(mdates.ConciseDateFormatter(axes[-1].xaxis.get_major_locator()))
    fig.text(0.01, 0.008, "İnce çizgiler aylık değerleri, kalın çizgiler 12 aylık hareketli ortalamayı gösterir.", fontsize=7.5, color="#455a64")
    fig.tight_layout()
    output = io.BytesIO()
    FigureCanvasAgg(fig).print_png(output)
    return output.getvalue()


def build_area_map_png(
    gdf: gpd.GeoDataFrame,
    point: tuple[float, float],
    *,
    source_note: str = "Kullanıcı çalışma alanı",
) -> bytes:
    fig = Figure(figsize=(10, 8), dpi=180, facecolor="#f4f2eb")
    ax = fig.add_subplot(111)
    geographic = gdf.to_crs(4326)
    map_crs = geographic.estimate_utm_crs() or "EPSG:6933"
    projected = geographic.to_crs(map_crs)
    geographic.plot(
        ax=ax,
        column=None,
        facecolor="#8fcf9a",
        edgecolor="#123b4a",
        linewidth=1.25,
    )
    area_km2 = projected.geometry.union_all().area / 1_000_000
    perimeter_km = projected.geometry.union_all().length / 1_000
    bounds = geographic.total_bounds
    span_x, span_y = bounds[2] - bounds[0], bounds[3] - bounds[1]
    ax.set_xlim(bounds[0] - span_x * 0.09, bounds[2] + span_x * 0.09)
    ax.set_ylim(bounds[1] - span_y * 0.13, bounds[3] + span_y * 0.09)
    ax.set_title("ÇALIŞMA ALANI SINIR HARİTASI", color="#092f3d", fontsize=16, weight="bold", pad=22)
    ax.text(0.5, 1.012, "Coğrafi gösterim ve temel alan özellikleri", transform=ax.transAxes, ha="center", va="bottom", fontsize=9, color="#48646d")
    ax.set_xlabel("Boylam (coğrafi derece)", labelpad=10)
    ax.set_ylabel("Enlem (coğrafi derece)", labelpad=10)
    ax.xaxis.set_major_formatter(FuncFormatter(lambda value, _: f"{abs(value):.2f}°{'D' if value >= 0 else 'B'}"))
    ax.yaxis.set_major_formatter(FuncFormatter(lambda value, _: f"{abs(value):.2f}°{'K' if value >= 0 else 'G'}"))
    ax.grid(color="#78909c", linestyle="--", linewidth=0.55, alpha=0.48)
    ax.set_axisbelow(True)
    centroid = geographic.dissolve().geometry.iloc[0].centroid
    ax.scatter([centroid.x], [centroid.y], marker="+", s=55, linewidth=1.4, color="#b71c1c", zorder=5)
    legend_handles = [
        Patch(facecolor="#8fcf9a", edgecolor="#123b4a", label="Çalışma / analiz alanı"),
        ax.plot([], [], marker="+", linestyle="none", color="#b71c1c", label="Geometrik merkez")[0],
    ]
    ax.legend(handles=legend_handles, loc="lower right", title="LEJANT", frameon=True, framealpha=1, edgecolor="#455a64", fontsize=8, title_fontsize=8.5)
    _add_north_arrow(ax)
    _add_scale_bar(ax, projected=False)
    ax.text(
        0.02,
        0.985,
        f"Alan: {area_km2:,.2f} km²\nÇevre: {perimeter_km:,.2f} km",
        transform=ax.transAxes,
        fontsize=8.5,
        color="#063447",
        va="top",
        bbox={"facecolor": "white", "edgecolor": "#137f83", "alpha": 0.96, "boxstyle": "round,pad=0.30"},
    )
    for spine in ax.spines.values():
        spine.set_linewidth(1.2)
        spine.set_color("#263238")
    fig.text(0.02, 0.018, f"Veri kaynağı: {source_note}", fontsize=7.5, color="#455a64")
    fig.text(0.98, 0.018, "Pafta üretimi: Zetriklim", ha="right", fontsize=7.5, color="#455a64")
    fig.add_artist(Rectangle((0.008, 0.008), 0.984, 0.984, transform=fig.transFigure, fill=False, edgecolor="#263238", linewidth=1.1))
    fig.subplots_adjust(left=0.10, right=0.975, bottom=0.10, top=0.89)
    output = io.BytesIO()
    FigureCanvasAgg(fig).print_png(output)
    return output.getvalue()


def geodata_to_gpkg(gdf: gpd.GeoDataFrame, point: tuple[float, float]) -> bytes:
    with tempfile.TemporaryDirectory(prefix="zetriklim_gpkg_") as temp:
        path = Path(temp) / "zetriklim-cbs.gpkg"
        gdf.to_file(path, layer="calisma_alani", driver="GPKG")
        point_gdf = gpd.GeoDataFrame(
            {
                "ornek_id": [1],
                "aciklama": ["İklim verisi örnekleme noktası"],
                "csv_baglanti_alani": ["Örnek ID"],
            },
            geometry=gpd.points_from_xy([point[1]], [point[0]]),
            crs=4326,
        )
        point_gdf.to_file(path, layer="ornekleme_noktasi", driver="GPKG")
        return path.read_bytes()


def build_map_vector_files(
    gdf: gpd.GeoDataFrame,
    metadata: dict[str, object],
) -> tuple[bytes, bytes]:
    """Harita kapsamını ve üretim bilgisini GeoPackage ile Shapefile paketi olarak verir."""
    geographic = gdf.to_crs(4326)
    record_count = len(geographic)

    def repeated(key: str, limit: int = 254) -> list[str]:
        value = str(metadata.get(key) or "")[:limit]
        return [value] * record_count

    layer = gpd.GeoDataFrame(
        {
            "HARITA": repeated("label"),
            "ANALIZ": repeated("analysis"),
            "KAYNAK": repeated("source"),
            "VERISETI": repeated("dataset"),
            "DEGISKEN": repeated("variable"),
            "BAS_TARIH": repeated("start", 32),
            "BIT_TARIH": repeated("end", 32),
            "KALITE": repeated("quality", 64),
        },
        geometry=geographic.geometry.reset_index(drop=True),
        crs=4326,
    )
    with tempfile.TemporaryDirectory(prefix="zetriklim_map_vector_") as temp:
        temp_path = Path(temp)
        gpkg_path = temp_path / "analiz-haritasi.gpkg"
        layer.to_file(gpkg_path, layer="analiz_haritasi", driver="GPKG")
        gpkg_bytes = gpkg_path.read_bytes()

        shp_dir = temp_path / "shapefile"
        shp_dir.mkdir()
        shp_path = shp_dir / "analiz_haritasi.shp"
        layer.to_file(shp_path, driver="ESRI Shapefile", encoding="UTF-8")
        (shp_dir / "harita-metadata.json").write_text(
            json.dumps(metadata, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        shp_output = io.BytesIO()
        with zipfile.ZipFile(shp_output, "w", zipfile.ZIP_DEFLATED) as archive:
            for path in sorted(shp_dir.iterdir()):
                archive.write(path, arcname=path.name)
        return gpkg_bytes, shp_output.getvalue()


def build_complete_package(files: dict[str, bytes]) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, content in files.items():
            archive.writestr(name, content)
    return output.getvalue()
