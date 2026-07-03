"""Excel, CSV, grafik ve CBS çıktıları."""

from __future__ import annotations

import io
import tempfile
import zipfile
from pathlib import Path

import geopandas as gpd
import pandas as pd
import rasterio
from rasterio.io import MemoryFile
from rasterio.plot import plotting_extent
from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.figure import Figure
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


NAVY = "063447"
TEAL = "00A6A6"
PALE = "E2F4F2"
AMBER = "FFAD33"


def dataframe_to_csv(data: pd.DataFrame) -> bytes:
    return data.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def build_excel(
    data: pd.DataFrame,
    *,
    metadata_rows: list[tuple[str, object]],
    area_rows: list[tuple[str, object]],
    analysis_tables: dict[str, pd.DataFrame] | None = None,
) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        data.to_excel(writer, sheet_name="İklim Verisi", index=False)
        pd.DataFrame(metadata_rows, columns=["Alan", "Değer"]).to_excel(writer, sheet_name="Kaynak ve Yöntem", index=False)
        pd.DataFrame(area_rows, columns=["Alan Özelliği", "Değer"]).to_excel(writer, sheet_name="Çalışma Alanı", index=False)
        data.describe(include="all").transpose().reset_index().to_excel(writer, sheet_name="Veri Özeti", index=False)
        for name, table in (analysis_tables or {}).items():
            table.to_excel(writer, sheet_name=name[:31], index=False)

        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = PatternFill("solid", fgColor=NAVY)
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for column in sheet.columns:
                letter = column[0].column_letter
                width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
                sheet.column_dimensions[letter].width = max(width, 12)

        data_sheet = writer.book["İklim Verisi"]
        data_sheet.sheet_view.showGridLines = False
        for row in range(2, data_sheet.max_row + 1):
            if row % 2 == 0:
                for cell in data_sheet[row]:
                    cell.fill = PatternFill("solid", fgColor="F2FAF9")
        if data_sheet.max_column >= 2 and data_sheet.max_row >= 3:
            chart = LineChart()
            chart.title = "İklim Zaman Serisi"
            chart.style = 13
            chart.y_axis.title = "Değer"
            chart.x_axis.title = "Tarih"
            chart.add_data(
                Reference(data_sheet, min_col=5, max_col=min(data_sheet.max_column, 8), min_row=1, max_row=data_sheet.max_row),
                titles_from_data=True,
            )
            chart.set_categories(Reference(data_sheet, min_col=1, min_row=2, max_row=data_sheet.max_row))
            chart.height = 9
            chart.width = 18
            summary_sheet = writer.book["Veri Özeti"]
            summary_sheet.add_chart(chart, "H2")
    return output.getvalue()


def build_raster_png(
    geotiff: bytes,
    title: str,
    *,
    boundary: gpd.GeoDataFrame | None = None,
    palette: str = "RdBu",
    colorbar_label: str = "SPI",
    fixed_range: tuple[float, float] | None = (-2.5, 2.5),
) -> bytes:
    with MemoryFile(geotiff) as memory:
        with memory.open() as src:
            raster = src.read(1, masked=True).astype(float).filled(float("nan"))
            extent = plotting_extent(src)
            raster_crs = src.crs
    fig = Figure(figsize=(9, 7), dpi=150, facecolor="#f7f5ee")
    ax = fig.add_subplot(111)
    range_args = (
        {"vmin": fixed_range[0], "vmax": fixed_range[1]}
        if fixed_range is not None
        else {}
    )
    image = ax.imshow(raster, cmap=palette, extent=extent, origin="upper", **range_args)
    if boundary is not None and raster_crs is not None:
        boundary.to_crs(raster_crs).boundary.plot(
            ax=ax,
            color="#082f49",
            linewidth=1.4,
            zorder=3,
        )
    ax.set_title(title, color="#063447", fontsize=13, weight="bold")
    ax.set_axis_off()
    colorbar = fig.colorbar(image, ax=ax, shrink=0.78)
    colorbar.set_label(colorbar_label)
    fig.tight_layout()
    output = io.BytesIO()
    FigureCanvasAgg(fig).print_png(output)
    return output.getvalue()


def build_timeseries_png(data: pd.DataFrame) -> bytes:
    numeric = data.drop(columns=["Örnek ID", "Enlem", "Boylam"], errors="ignore").select_dtypes("number")
    fig = Figure(figsize=(12, 6), dpi=150, facecolor="#f7f5ee")
    ax = fig.add_subplot(111)
    for column in numeric.columns[:5]:
        ax.plot(data["Tarih"], numeric[column], linewidth=1.1, label=column)
    ax.set_title("Zetriklim – Analiz Zaman Serisi", color="#063447", fontsize=14, weight="bold")
    ax.set_xlabel("Tarih")
    ax.grid(alpha=0.20)
    ax.legend(loc="best", fontsize=8)
    fig.tight_layout()
    output = io.BytesIO()
    FigureCanvasAgg(fig).print_png(output)
    return output.getvalue()


def build_area_map_png(gdf: gpd.GeoDataFrame, point: tuple[float, float]) -> bytes:
    fig = Figure(figsize=(9, 7), dpi=150, facecolor="#f7f5ee")
    ax = fig.add_subplot(111)
    projected = gdf.to_crs(3857)
    projected.plot(ax=ax, facecolor="#62d5cc", edgecolor="#063447", linewidth=1.6)
    area_km2 = projected.geometry.union_all().area / 1_000_000
    ax.set_title("Çalışma Alanı / Havza Sınırı", color="#063447", fontsize=13, weight="bold")
    ax.text(
        0.02,
        0.02,
        f"Alan: {area_km2:,.2f} km²",
        transform=ax.transAxes,
        fontsize=9,
        color="#063447",
        bbox={"facecolor": "white", "edgecolor": "#00a6a6", "alpha": 0.9, "boxstyle": "round,pad=0.4"},
    )
    ax.set_axis_off()
    fig.tight_layout()
    output = io.BytesIO()
    FigureCanvasAgg(fig).print_png(output)
    return output.getvalue()


def geodata_to_gpkg(gdf: gpd.GeoDataFrame, point: tuple[float, float]) -> bytes:
    with tempfile.TemporaryDirectory(prefix="zetriklim_gpkg_") as temp:
        path = Path(temp) / "zetriklim-cbs.gpkg"
        gdf.to_file(path, layer="calisma_alani", driver="GPKG")
        point_gdf = gpd.GeoDataFrame(
            {
                "ornek_id": [1],
                "aciklama": ["İklim verisi örnekleme noktası"],
                "csv_baglanti_alani": ["Örnek ID"],
            },
            geometry=gpd.points_from_xy([point[1]], [point[0]]),
            crs=4326,
        )
        point_gdf.to_file(path, layer="ornekleme_noktasi", driver="GPKG")
        return path.read_bytes()


def build_complete_package(files: dict[str, bytes]) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, content in files.items():
            archive.writestr(name, content)
    return output.getvalue()
