"""Excel, CSV, grafik ve CBS çıktıları."""

from __future__ import annotations

import io
import tempfile
import zipfile
from pathlib import Path

import geopandas as gpd
import numpy as np
import pandas as pd
import rasterio
from rasterio.io import MemoryFile
from rasterio.plot import plotting_extent
from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.figure import Figure
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


NAVY = "063447"
TEAL = "00A6A6"
PALE = "E2F4F2"
AMBER = "FFAD33"


def dataframe_to_csv(data: pd.DataFrame) -> bytes:
    return data.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def build_excel(
    data: pd.DataFrame,
    *,
    metadata_rows: list[tuple[str, object]],
    area_rows: list[tuple[str, object]],
    analysis_tables: dict[str, pd.DataFrame] | None = None,
) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        data.to_excel(writer, sheet_name="İklim Verisi", index=False)
        pd.DataFrame(metadata_rows, columns=["Alan", "Değer"]).to_excel(writer, sheet_name="Kaynak ve Yöntem", index=False)
        pd.DataFrame(area_rows, columns=["Alan Özelliği", "Değer"]).to_excel(writer, sheet_name="Çalışma Alanı", index=False)
        data.describe(include="all").transpose().reset_index().to_excel(writer, sheet_name="Veri Özeti", index=False)
        for name, table in (analysis_tables or {}).items():
            table.to_excel(writer, sheet_name=name[:31], index=False)

        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = PatternFill("solid", fgColor=NAVY)
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for column in sheet.columns:
                letter = column[0].column_letter
                width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
                sheet.column_dimensions[letter].width = max(width, 12)

        data_sheet = writer.book["İklim Verisi"]
        data_sheet.sheet_view.showGridLines = False
        for row in range(2, data_sheet.max_row + 1):
            if row % 2 == 0:
                for cell in data_sheet[row]:
                    cell.fill = PatternFill("solid", fgColor="F2FAF9")
        if data_sheet.max_column >= 2 and data_sheet.max_row >= 3:
            chart = LineChart()
            chart.title = "İklim Zaman Serisi"
            chart.style = 13
            chart.y_axis.title = "Değer"
            chart.x_axis.title = "Tarih"
            chart.add_data(
                Reference(data_sheet, min_col=5, max_col=min(data_sheet.max_column, 8), min_row=1, max_row=data_sheet.max_row),
                titles_from_data=True,
            )
            chart.set_categories(Reference(data_sheet, min_col=1, min_row=2, max_row=data_sheet.max_row))
            chart.height = 9
            chart.width = 18
            summary_sheet = writer.book["Veri Özeti"]
            summary_sheet.add_chart(chart, "H2")
    return output.getvalue()


def build_raster_png(
    geotiff: bytes,
    title: str,
    *,
    boundary: gpd.GeoDataFrame | None = None,
    palette: str = "RdBu",
    colorbar_label: str = "SPI",
    fixed_range: tuple[float, float] | None = (-2.5, 2.5),
) -> bytes:
    with MemoryFile(geotiff) as memory:
        with memory.open() as src:
            raster = src.read(1, masked=True).astype(float).filled(float("nan"))
            extent = plotting_extent(src)
            raster_crs = src.crs
    fig = Figure(figsize=(9, 7), dpi=150, facecolor="#f7f5ee")
    ax = fig.add_subplot(111)
    range_args = (
        {"vmin": fixed_range[0], "vmax": fixed_range[1]}
        if fixed_range is not None
        else {}
    )
    image = ax.imshow(raster, cmap=palette, extent=extent, origin="upper", **range_args)
    if boundary is not None and raster_crs is not None:
        boundary.to_crs(raster_crs).boundary.plot(
            ax=ax,
            color="#082f49",
            linewidth=1.4,
            zorder=3,
        )
    ax.set_title(title, color="#063447", fontsize=13, weight="bold")
    ax.set_axis_off()
    colorbar = fig.colorbar(image, ax=ax, shrink=0.78)
    colorbar.set_label(colorbar_label)
    fig.tight_layout()
    output = io.BytesIO()
    FigureCanvasAgg(fig).print_png(output)
    return output.getvalue()


def build_timeseries_png(data: pd.DataFrame) -> bytes:
    frame = data.copy()
    frame["Tarih"] = pd.to_datetime(frame["Tarih"], errors="coerce")
    frame = frame.dropna(subset=["Tarih"]).sort_values("Tarih")
    numeric = frame.drop(
        columns=["Örnek ID", "Enlem", "Boylam"], errors="ignore"
    ).select_dtypes("number")
    fig = Figure(figsize=(13, 7), dpi=160, facecolor="#f7f5ee")
    ax = fig.add_subplot(111, facecolor="#fffdf8")

    ndvi_column = next(
        (column for column in numeric.columns if str(column).strip().upper() == "NDVI"),
        None,
    )
    if ndvi_column:
        series = pd.to_numeric(numeric[ndvi_column], errors="coerce")
        series = series.where(series.between(-1.0, 1.0))
        plotted = (
            pd.DataFrame({"Tarih": frame["Tarih"], "NDVI": series})
            .groupby("Tarih", as_index=False)["NDVI"]
            .mean()
        )
        bands = [
            (-1.0, 0.0, "#c7dcef", "Su, kar veya gölge (−1–0)"),
            (0.0, 0.2, "#d9c7a3", "Çıplak yüzey / yapılaşma (0–0,2)"),
            (0.2, 0.4, "#f0df78", "Seyrek bitki örtüsü (0,2–0,4)"),
            (0.4, 0.6, "#8fcf72", "Orta yoğunlukta bitki (0,4–0,6)"),
            (0.6, 1.0, "#2f8f4e", "Yoğun ve sağlıklı bitki (0,6–1)"),
        ]
        for lower, upper, color, label in bands:
            ax.axhspan(lower, upper, color=color, alpha=0.24, label=label, zorder=0)
        ax.axhline(0, color="#596b73", linewidth=0.8, alpha=0.65)
        ax.plot(
            plotted["Tarih"],
            plotted["NDVI"],
            color="#075b68",
            linewidth=1.5,
            marker="o",
            markersize=2.4,
            markevery=max(len(plotted) // 140, 1),
            label="Havza ortalama NDVI",
            zorder=3,
        )
        if plotted["NDVI"].notna().sum() >= 5:
            rolling = plotted["NDVI"].rolling(5, center=True, min_periods=3).median()
            ax.plot(
                plotted["Tarih"],
                rolling,
                color="#ef6c57",
                linewidth=2.0,
                label="5 gözlem hareketli medyan",
                zorder=4,
            )
        valid_count = int(plotted["NDVI"].notna().sum())
        missing_count = int(plotted["NDVI"].isna().sum())
        ax.set_ylim(-1.02, 1.02)
        ax.set_ylabel("NDVI (birimsiz; −1 ile +1)")
        ax.set_title(
            "NDVI Zaman Serisi – Havza Alan Ortalaması",
            color="#063447",
            fontsize=15,
            weight="bold",
        )
        ax.text(
            0.01,
            0.02,
            f"Geçerli gözlem: {valid_count:,}  ·  NoData/elenen: {missing_count:,}",
            transform=ax.transAxes,
            fontsize=8.5,
            color="#315b64",
            bbox={
                "facecolor": "white",
                "edgecolor": "#8abfbb",
                "alpha": 0.92,
                "boxstyle": "round,pad=0.35",
            },
        )
        ax.legend(
            loc="upper left",
            bbox_to_anchor=(1.01, 1.0),
            borderaxespad=0,
            fontsize=8,
            title="NDVI açıklaması",
            title_fontsize=9,
            frameon=True,
        )
    else:
        colors = ["#075b68", "#ef6c57", "#7b61a8", "#d18b00", "#277da1"]
        for index, column in enumerate(numeric.columns[:5]):
            values = pd.to_numeric(numeric[column], errors="coerce").replace(
                [np.inf, -np.inf], np.nan
            )
            values = values.mask(values <= -9990)
            ax.plot(
                frame["Tarih"],
                values,
                linewidth=1.35,
                color=colors[index % len(colors)],
                label=str(column),
            )
        ax.set_title(
            "Zetriklim – Analiz Zaman Serisi",
            color="#063447",
            fontsize=14,
            weight="bold",
        )
        ax.set_ylabel("Değer")
        ax.legend(loc="best", fontsize=8, title="Değişken")
    ax.set_xlabel("Tarih")
    ax.grid(axis="both", alpha=0.22, linestyle="--", linewidth=0.6)
    fig.tight_layout(rect=(0, 0, 0.75, 1) if ndvi_column else None)
    output = io.BytesIO()
    FigureCanvasAgg(fig).print_png(output)
    return output.getvalue()


def build_area_map_png(gdf: gpd.GeoDataFrame, point: tuple[float, float]) -> bytes:
    fig = Figure(figsize=(9, 7), dpi=150, facecolor="#f7f5ee")
    ax = fig.add_subplot(111)
    projected = gdf.to_crs(3857)
    projected.plot(ax=ax, facecolor="#62d5cc", edgecolor="#063447", linewidth=1.6)
    area_km2 = projected.geometry.union_all().area / 1_000_000
    ax.set_title("Çalışma Alanı / Havza Sınırı", color="#063447", fontsize=13, weight="bold")
    ax.text(
        0.02,
        0.02,
        f"Alan: {area_km2:,.2f} km²",
        transform=ax.transAxes,
        fontsize=9,
        color="#063447",
        bbox={"facecolor": "white", "edgecolor": "#00a6a6", "alpha": 0.9, "boxstyle": "round,pad=0.4"},
    )
    ax.set_axis_off()
    fig.tight_layout()
    output = io.BytesIO()
    FigureCanvasAgg(fig).print_png(output)
    return output.getvalue()


def geodata_to_gpkg(gdf: gpd.GeoDataFrame, point: tuple[float, float]) -> bytes:
    with tempfile.TemporaryDirectory(prefix="zetriklim_gpkg_") as temp:
        path = Path(temp) / "zetriklim-cbs.gpkg"
        gdf.to_file(path, layer="calisma_alani", driver="GPKG")
        point_gdf = gpd.GeoDataFrame(
            {
                "ornek_id": [1],
                "aciklama": ["İklim verisi örnekleme noktası"],
                "csv_baglanti_alani": ["Örnek ID"],
            },
            geometry=gpd.points_from_xy([point[1]], [point[0]]),
            crs=4326,
        )
        point_gdf.to_file(path, layer="ornekleme_noktasi", driver="GPKG")
        return path.read_bytes()


def build_complete_package(files: dict[str, bytes]) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, content in files.items():
            archive.writestr(name, content)
    return output.getvalue()
